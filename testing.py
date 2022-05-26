#Importing Liabraries
import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
import tkinter.messagebox as tmsg

#Creating Interface
root=Tk()
Label(text="Mobile Cafe",font="comicsansms 11 bold",pady=15,padx=25).grid(row=1,column=1)
root.geometry("677x434")
root.title("Point of Sale")
Button(root, text="Create File", command=lambda: newfile()).grid(row=1, column=2)
Button(root, text="Load File", command=lambda: loadfile()).grid(row=3, column=2)

file = StringVar()
def newfile():
    Label(root, text="File Name: ", font="comicsansms 11").grid(row=1, column=3)
    file_entry = Entry(root, textvariable=file).grid(row=1, column=4)
    Button(root, text="Enter", command=lambda: (wb.save(file.get() + ".xlsx"),tmsg.showinfo("File Info","File created!"))).grid(row=1, column=5)

fileload = StringVar()
def loadfile():
    Label(root, text="Load File Name: ", font="comicsansms 11").grid(row=3, column=3)
    fileload_entry = Entry(root, textvariable=fileload).grid(row=3, column=4)
    Button(root, text="Enter", command=lambda: (flload() , tmsg.showinfo("File Info","File Loaded!"))).grid(row=3, column=5)


wb = Workbook()

ws = wb.active
ws['A1'] = "Name of Product:"
ws['B1'] = "Qty:"
ws['C1'] = "Total Price:"
ws['D1'] = "Product Key"
ws['E1'] = "Per Item Price:"
ws['F1'] = "Date and Time"




def flload():
    wb2 = load_workbook(fileload.get() + ".xlsx")

    ws = wb2.worksheets[0]

    Label(text="Data Entry", font="comicsansms 11 bold", pady=15, padx=25).grid(row=6, column=2)

    Label(text="Name of Product", font="comicsansms 11").grid(row=7, column=1)
    Label(text="Qty", font="comicsansms 11").grid(row=8, column=1)
    Label(text="Total Price", font="comicsansms 11").grid(row=9, column=1)

    Label(text="Data Call", font="comicsansms 11 bold", pady=15, padx=25).grid(row=11, column=2)
    Label(text="Name of Product", font="comicsansms 11").grid(row=12, column=1)
    Label(text="Qty", font="comicsansms 11").grid(row=13, column=1)
    nameinit = StringVar()
    qtyinit = IntVar()
    name_entry = Entry(root, textvariable=nameinit).grid(row=12, column=2)
    qty_entry = Entry(root, textvariable=qtyinit).grid(row=13, column=2)
    Button(root, text="Enter", command=lambda: pts()).grid(row=14, column=2)

    def pts():
        Label(text="Data Call", font="comicsansms 11 bold").grid(row=11, column=2)
        fint = load_workbook(fileload.get() + ".xlsx")
        filed = fint.active
        named = fint["Sheet"]
        sar = nameinit.get()
        for col in named.rows:
            if (col[0].value == sar):
                founded = sar

        rval = pd.read_excel(fileload.get() + ".xlsx", sheet_name="Sheet")
        rval.set_index("Name of Product:", inplace=True)
        rval2 = rval["Per Item Price:"][founded]
        rval3 = qtyinit.get()
        tval = int(rval3) * int(rval2)
        print("Products:" + "    " + "Qty:" + "     " + "Price:" + "     " + "Total:" + "\n" + "    " + founded + "       ",rval3, "     ", rval2, "     ", tval)


    name = StringVar()
    qty = IntVar()
    price = IntVar()

    name_entry = Entry(root, textvariable=name).grid(row=7, column=2)
    qty_entry = Entry(root, textvariable=qty).grid(row=8, column=2)
    price_entry = Entry(root, textvariable=price).grid(row=9, column=2)
    Button(root, text="Enter", command=lambda: data()).grid(row=10, column=2)
    def data():
        ws.append([name.get(),
                   qty.get(),
                   price.get(),
                   0,
                   price.get()/qty.get(),
                   datetime.datetime.now(),
                   ])
        wb2.save(fileload.get() + ".xlsx")

root.mainloop()
exit()
